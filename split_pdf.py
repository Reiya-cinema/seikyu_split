from __future__ import annotations

import argparse
import io
import os
import re
import sys
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pdfplumber
import pymupdf
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def mm_to_pt(mm: float) -> float:
    return float(mm) * 72.0 / 25.4


def _normalize_text(text: str | None) -> str:
    if text is None:
        return ""
    return str(text).strip()


def _normalize_for_match(text: str | None) -> str:
    normalized = _normalize_text(text)
    compact = re.sub(r"[\s　]+", "", normalized)
    return re.sub(r"(.)\1+", r"\1", compact)


def _contains_garbled_text(text: str) -> bool:
    return "(cid:" in text


def _parse_rect_pt(value: str | None) -> tuple[float, float, float, float] | None:
    text = _normalize_text(value)
    if not text:
        return None
    nums = re.findall(r"-?\d+(?:\.\d+)?", text)
    if len(nums) < 4:
        return None
    x, y, w, h = (float(nums[0]), float(nums[1]), float(nums[2]), float(nums[3]))
    return x, y, w, h


def _rect_pt_to_pdf_bbox(rect_pt: tuple[float, float, float, float]) -> tuple[float, float, float, float]:
    x, y, w, h = rect_pt
    x0 = x
    y0 = y
    x1 = x + w
    y1 = y + h
    return x0, y0, x1, y1


def _bbox_within_page(bbox: tuple[float, float, float, float], page: pdfplumber.page.Page) -> bool:
    x0, y0, x1, y1 = bbox
    return x0 >= 0 and y0 >= 0 and x1 <= page.width and y1 <= page.height and x1 > x0 and y1 > y0


def _resolve_bbox(page: pdfplumber.page.Page, rect: tuple[float, float, float, float]) -> tuple[float, float, float, float] | None:
    bbox = _rect_pt_to_pdf_bbox(rect)
    if _bbox_within_page(bbox, page):
        return bbox
    return None


class OcrExtractor:
    def __init__(self, pdf_path: Path):
        self.pdf_doc = pymupdf.open(pdf_path)
        self.language = self._resolve_ocr_language()

    def _resolve_ocr_language(self) -> str:
        requested = os.getenv("OCR_LANG", "jpn")
        requested_parts = [part for part in requested.split("+") if part]
        try:
            available = set(pytesseract.get_languages(config=""))
        except Exception:
            available = set()

        if requested_parts and all(part in available for part in requested_parts):
            return requested
        if "jpn" in available:
            return "jpn"
        if "eng" in available:
            return "eng"
        if available:
            return sorted(available)[0]
        return requested

    def close(self) -> None:
        self.pdf_doc.close()

    def extract_text(self, page_index: int, bbox: tuple[float, float, float, float], mode: str = "general") -> str:
        page = self.pdf_doc[page_index]
        rect = pymupdf.Rect(*bbox)
        pix = page.get_pixmap(matrix=pymupdf.Matrix(3, 3), clip=rect, alpha=False)
        image_bytes = pix.tobytes("png")
        if mode == "keyword":
            ocr_config = "--oem 1 --psm 7"
        else:
            ocr_config = "--oem 1 --psm 6"
        with Image.open(io.BytesIO(image_bytes)) as image:
            text = pytesseract.image_to_string(image, lang=self.language, config=ocr_config)
        return " ".join((text or "").split())


def _extract_text_in_rect(
    page: pdfplumber.page.Page,
    rect_pt: tuple[float, float, float, float] | None,
    page_index: int,
    ocr_extractor: OcrExtractor | None,
    ocr_mode: str = "general",
) -> str:
    if rect_pt is None:
        return ""
    bbox = _resolve_bbox(page, rect_pt)
    if bbox is None:
        return ""
    text = page.crop(bbox).extract_text() or ""
    normalized_text = " ".join(text.split())
    if ocr_extractor is None:
        return normalized_text
    if normalized_text and not _contains_garbled_text(normalized_text):
        return normalized_text

    ocr_text = ocr_extractor.extract_text(page_index, bbox, mode=ocr_mode)
    return ocr_text or normalized_text


def _resolve_dir(base: Path, raw: str) -> Path:
    path = Path(raw)
    if path.is_absolute():
        return path
    return (base / path).resolve()


@dataclass
class LayoutRule:
    layout_name: str
    judge_rect_pt: tuple[float, float, float, float] | None
    judge_keyword: str
    addressee_rect_pt: tuple[float, float, float, float] | None


@dataclass
class RenameRule:
    keyword: str
    output_name: str


@dataclass
class AppSettings:
    input_dir: Path
    layout_rules: list[LayoutRule]
    rename_rules: list[RenameRule]


def _iter_rows_from(sheet: Worksheet, start_row: int) -> Iterable[tuple[int, list[object]]]:
    for row in range(start_row, sheet.max_row + 1):
        yield row, [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]


def load_settings(ws_settings: Worksheet, excel_dir: Path) -> AppSettings:
    input_dir_raw = "input"
    layout_rules: list[LayoutRule] = []
    rename_rules: list[RenameRule] = []

    for _, row in _iter_rows_from(ws_settings, 4):
        judge_rect = _parse_rect_pt(row[0] if len(row) > 0 else None)
        judge_text = _normalize_text(row[1] if len(row) > 1 else None)
        layout_name = _normalize_text(row[2] if len(row) > 2 else None)
        addressee_rect = _parse_rect_pt(row[3] if len(row) > 3 else None)

        if judge_rect or judge_text or layout_name or addressee_rect:
            layout_rules.append(
                LayoutRule(
                    layout_name=layout_name or "未設定",
                    judge_rect_pt=judge_rect,
                    judge_keyword=judge_text,
                    addressee_rect_pt=addressee_rect,
                )
            )

        rename_keyword = _normalize_text(row[5] if len(row) > 5 else None)
        rename_output = _normalize_text(row[6] if len(row) > 6 else None)
        if rename_keyword and rename_output:
            rename_rules.append(RenameRule(keyword=rename_keyword, output_name=rename_output))

        setting_key = _normalize_text(row[8] if len(row) > 8 else None)
        setting_value = _normalize_text(row[9] if len(row) > 9 else None)
        if setting_key == "入力ファイル置き場(フォルダ名)" and setting_value:
            input_dir_raw = setting_value

    return AppSettings(
        input_dir=_resolve_dir(excel_dir, input_dir_raw),
        layout_rules=layout_rules,
        rename_rules=rename_rules,
    )


def _pick_layout(
    page: pdfplumber.page.Page,
    page_index: int,
    rules: list[LayoutRule],
    ocr_extractor: OcrExtractor | None,
) -> tuple[str, str, tuple[float, float, float, float] | None]:
    for rule in rules:
        judged_text = _extract_text_in_rect(page, rule.judge_rect_pt, page_index, ocr_extractor, ocr_mode="keyword")
        judged_text_compact = _normalize_for_match(judged_text)
        judge_keyword_compact = _normalize_for_match(rule.judge_keyword)
        if not rule.judge_keyword:
            if judged_text:
                return rule.layout_name, judged_text, rule.addressee_rect_pt
            continue

        if judge_keyword_compact and judge_keyword_compact in judged_text_compact:
            return rule.layout_name, judged_text, rule.addressee_rect_pt

    if rules:
        first_rule = rules[0]
        return first_rule.layout_name, "", first_rule.addressee_rect_pt

    return "未判定", "", None


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]", "_", name).strip()
    return cleaned[:120] if cleaned else "未設定"


def _build_output_candidate(addressee: str, rename_rules: list[RenameRule], page_no: int) -> str:
    for rule in rename_rules:
        if rule.keyword and rule.keyword in addressee:
            return _sanitize_filename(rule.output_name)
    if addressee:
        return _sanitize_filename(addressee)
    return f"page_{page_no:04d}"


def run_scan_mode(excel_path: Path) -> None:
    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

    try:
        wb = load_workbook(excel_path, keep_vba=True)
    except PermissionError as exc:
        raise PermissionError(f"Excelファイルにアクセスできません（使用中/読み取り専用の可能性）: {excel_path}") from exc

    try:
        if "設定" not in wb.sheetnames:
            raise ValueError("設定シートが見つかりません。")

        ws_settings = wb["設定"]
        settings = load_settings(ws_settings, excel_path.parent)

        pdf_candidates = sorted([*settings.input_dir.glob("*.pdf"), *settings.input_dir.glob("*.PDF")])
        if not pdf_candidates:
            raise FileNotFoundError(f"入力フォルダにPDFがありません: {settings.input_dir}")

        target_pdf = pdf_candidates[0]
        ocr_extractor: OcrExtractor | None = None
        try:
            ocr_extractor = OcrExtractor(target_pdf)
        except Exception:
            ocr_extractor = None
        results: list[tuple[int, str, str, str, str]] = []
        prev_key: tuple[str, str] | None = None

        try:
            with pdfplumber.open(target_pdf) as pdf:
                for index, page in enumerate(pdf.pages, start=1):
                    layout_name, judged_text, addressee_rect = _pick_layout(page, index - 1, settings.layout_rules, ocr_extractor)
                    addressee_text = _extract_text_in_rect(page, addressee_rect, index - 1, ocr_extractor, ocr_mode="general")
                    output_name = _build_output_candidate(addressee_text, settings.rename_rules, index)

                    current_key = (layout_name, addressee_text)
                    merge_flag = "結合" if prev_key == current_key else "新規"
                    prev_key = current_key

                    results.append((index, layout_name, addressee_text or judged_text, output_name, merge_flag))
        finally:
            if ocr_extractor is not None:
                ocr_extractor.close()

        out_sheet_name = "ワークシート" if "ワークシート" in wb.sheetnames else "実行シート"
        if out_sheet_name not in wb.sheetnames:
            raise ValueError("結果書き込み先シート（ワークシート/実行シート）が見つかりません。")
        ws_out = wb[out_sheet_name]

        for row in range(5, ws_out.max_row + 1):
            for col in range(2, 7):
                ws_out.cell(row, col).value = None

        ws_out.cell(5, 2).value = target_pdf.name
        ws_out.cell(7, 2).value = "ページ数"
        ws_out.cell(7, 3).value = "レイアウト判断"
        ws_out.cell(7, 4).value = "キーワード"
        ws_out.cell(7, 5).value = "出力ファイル名"
        ws_out.cell(7, 6).value = "結合フラグ"

        write_row = 8
        for page_no, layout_name, keyword, output_name, merge_flag in results:
            ws_out.cell(write_row, 2).value = page_no
            ws_out.cell(write_row, 3).value = layout_name
            ws_out.cell(write_row, 4).value = keyword
            ws_out.cell(write_row, 5).value = output_name
            ws_out.cell(write_row, 6).value = merge_flag
            write_row += 1

        try:
            wb.save(excel_path)
        except PermissionError as exc:
            raise PermissionError(f"Excelファイルの保存に失敗しました（使用中/読み取り専用の可能性）: {excel_path}") from exc

        print(f"scan完了: {target_pdf.name} ({len(results)}ページ)")
    finally:
        _close_workbook_safely(wb)


def run_execute_mode(excel_path: Path) -> None:
    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

    wb = None
    try:
        wb = load_workbook(excel_path, data_only=True, keep_vba=True)
    except PermissionError as exc:
        raise PermissionError(f"Excelファイルにアクセスできません（使用中/読み取り専用の可能性）: {excel_path}") from exc

    try:
        if "設定" not in wb.sheetnames:
            raise ValueError("設定シートが見つかりません。")

        ws_settings = wb["設定"]
        settings = load_settings(ws_settings, excel_path.parent)

        pdf_candidates = sorted([*settings.input_dir.glob("*.pdf"), *settings.input_dir.glob("*.PDF")])
        if not pdf_candidates:
            raise FileNotFoundError(f"入力フォルダにPDFがありません: {settings.input_dir}")
        target_pdf_path = pdf_candidates[0]

        ws_name = "ワークシート" if "ワークシート" in wb.sheetnames else "実行シート"
        if ws_name not in wb.sheetnames:
            raise ValueError(f"指定されたシートが見つかりません: {ws_name}")
        ws = wb[ws_name]

        page_groups: list[tuple[str, list[int]]] = []
        current_name: str | None = None
        current_pages: list[int] = []

        # 8行目からデータ開始
        for row in range(8, ws.max_row + 1):
            cell_val = ws.cell(row, 2).value
            name_val = ws.cell(row, 5).value
            
            if cell_val is None:
                continue
            try:
                page_no = int(cell_val)
            except (ValueError, TypeError):
                continue

            out_name = str(name_val).strip() if name_val else "不明"
            
            if current_name == out_name:
                current_pages.append(page_no)
            else:
                if current_name is not None and current_pages:
                    page_groups.append((current_name, list(current_pages)))
                
                current_name = out_name
                current_pages = [page_no]

        if current_name is not None and current_pages:
            page_groups.append((current_name, list(current_pages)))

        if not page_groups:
            print("出力対象のページが見つかりませんでした。")
            return

        output_dir = excel_path.parent / "output"
        output_dir.mkdir(exist_ok=True)
        
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        save_dir = output_dir / timestamp
        save_dir.mkdir(exist_ok=True)

        try:
            doc = pymupdf.open(target_pdf_path)
            total_pages = len(doc)
            
            for out_name, pages in page_groups:
                new_doc = pymupdf.open()
                valid_pages = [p - 1 for p in pages if 1 <= p <= total_pages]
                
                if not valid_pages:
                    new_doc.close()
                    continue

                new_doc.insert_pdf(doc, from_page=valid_pages[0], to_page=valid_pages[0])
                for p_idx in valid_pages[1:]:
                    new_doc.insert_pdf(doc, from_page=p_idx, to_page=p_idx)
                
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', out_name)
                if not safe_name.lower().endswith(".pdf"):
                    safe_name += ".pdf"
                
                save_path = save_dir / safe_name
                new_doc.save(save_path)
                new_doc.close()
            
            doc.close()
        except Exception as e:
             raise RuntimeError(f"PDF処理中にエラーが発生しました: {e}")

        print(f"出力完了: {len(page_groups)}ファイルを作成しました。\n保存先: {save_dir}")

    finally:
        if wb:
            _close_workbook_safely(wb)


def run_execute_mode_with_popup(excel_path: Path) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox
    except Exception:
        run_execute_mode(excel_path)
        return

    status: dict[str, Exception | None] = {"error": None}

    def worker() -> None:
        try:
            run_execute_mode(excel_path)
        except Exception as exc:
            status["error"] = exc

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    root = tk.Tk()
    root.title("請求書分割 - 出力実行")
    root.resizable(False, False)
    root.geometry("400x120")

    message = tk.StringVar(value="PDF分割・保存中です...")
    elapsed = tk.StringVar(value="経過: 0秒")

    tk.Label(root, textvariable=message, padx=16, pady=16).pack()
    tk.Label(root, textvariable=elapsed, padx=16, pady=4).pack()

    started_at = time.time()

    def block_close() -> None:
        return

    root.protocol("WM_DELETE_WINDOW", block_close)

    def poll() -> None:
        elapsed.set(f"経過: {int(time.time() - started_at)}秒")
        if thread.is_alive():
            root.after(200, poll)
            return

        error = status["error"]
        if error is None:
            messagebox.showinfo("完了", "出力処理が完了しました。", parent=root)
        else:
            messagebox.showerror("エラー", f"出力処理に失敗しました。\n{error}", parent=root)
        root.destroy()

    root.after(200, poll)
    root.mainloop()

    if status["error"] is not None:
        raise status["error"]


def _close_workbook_safely(wb) -> None:
    # VBAプロジェクトが含まれている場合の安全なクローズ処理
    vba_archive = getattr(wb, "vba_archive", None)
    if vba_archive is not None:
        try:
            archive_fp = getattr(vba_archive, "fp", None)
            if archive_fp is not None and not archive_fp.closed:
                vba_archive.close()
        except Exception:
            pass
        try:
            vba_archive.fp = None
        except Exception:
            pass
    wb.close()


def run_scan_mode_with_popup(excel_path: Path) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox
    except Exception:
        run_scan_mode(excel_path)
        return

    status: dict[str, Exception | None] = {"error": None}

    def worker() -> None:
        try:
            run_scan_mode(excel_path)
        except Exception as exc:
            status["error"] = exc

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    root = tk.Tk()
    root.title("請求書分割")
    root.resizable(False, False)
    root.geometry("360x120")

    message = tk.StringVar(value="処理中です。しばらくお待ちください...")
    elapsed = tk.StringVar(value="経過: 0秒")

    tk.Label(root, textvariable=message, padx=16, pady=16).pack()
    tk.Label(root, textvariable=elapsed, padx=16, pady=4).pack()

    started_at = time.time()

    def block_close() -> None:
        return

    root.protocol("WM_DELETE_WINDOW", block_close)

    def poll() -> None:
        elapsed.set(f"経過: {int(time.time() - started_at)}秒")
        if thread.is_alive():
            root.after(200, poll)
            return

        error = status["error"]
        if error is None:
            messagebox.showinfo("完了", "スキャン処理が完了しました。", parent=root)
        else:
            messagebox.showerror("エラー", f"スキャン処理に失敗しました。\n{error}", parent=root)
        root.destroy()

    root.after(200, poll)
    root.mainloop()

    if status["error"] is not None:
        raise status["error"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="請求書PDF分割ツール")
    parser.add_argument("--mode", required=True, choices=["scan", "execute"], help="実行モード")
    parser.add_argument("--excel", required=True, help="Excelファイルのフルパス")
    parser.add_argument("--popup", action="store_true", help="処理中ポップアップを表示する")
    return parser.parse_args()


def main() -> int:
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8-sig")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8-sig")
    except Exception:
        pass

    args = parse_args()
    excel_path = Path(args.excel)

    try:
        if args.mode == "scan":
            if args.popup:
                run_scan_mode_with_popup(excel_path)
            else:
                run_scan_mode(excel_path)
            return 0

        if args.mode == "execute":
            if args.popup:
                run_execute_mode_with_popup(excel_path)
            else:
                run_execute_mode(excel_path)
            return 0

        print("不明なモードです。")
        return 1
    except Exception as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())